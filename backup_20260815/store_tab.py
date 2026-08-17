from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QFileDialog, QMessageBox, QInputDialog, QLabel,
                             QApplication, QComboBox, QGridLayout, QFrame,
                             QLineEdit, QAbstractItemView, QDialog, QProgressBar)
from PyQt6.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QImage
from table_widget import PromoTableWidget
from data_manager import DataManager
from promotion_manager import PromotionManager
from upload_dialog import UploadDialog
from competitor_manager import CompetitorManager
from multi_select_combo import MultiSelectComboBox
import os
import re
import pandas as pd


class StoreTab(QWidget):
    """单店铺标签页 - 带筛选行"""

    def __init__(self, store_name: str, parent=None):
        super().__init__(parent)
        self.store_name = store_name
        self.data_manager = DataManager(store_name)
        self.promo_manager = PromotionManager(self.data_manager)
        self.filter_widgets = {}
        self.color_filter_widget = None
        self.category_input = None
        self.filter_container = None
        self.copy_btn = None
        self.copy_status = None
        self.search_input = None
        self.search_results = []
        self.search_current_index = -1
        self.search_highlighted_row = -1
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self._do_search)
        self._pending_search_text = ""

        # ===== 设置默认路径 =====
        if store_name == "6号店":
            self.default_dir = r"C:\Users\liuzh\Desktop\Super Browser\李军WB 5-6号"
        elif store_name == "8号店":
            self.default_dir = r"C:\Users\liuzh\Desktop\Super Browser\本土WB 8号"
        else:
            self.default_dir = os.path.expanduser("~")

        self.setup_ui()
        self.load_data()

    def import_and_sort(self, file_path: str):
        """
        从Excel/CSV读取顺序，然后按照这个顺序重新排列当前数据
        """
        try:
            # 1. 读取用户整理好的顺序文件
            if file_path.endswith('.csv'):
                df_order = pd.read_csv(file_path, encoding='utf-8-sig', dtype=str, keep_default_na=False)
            else:
                df_order = pd.read_excel(file_path, dtype=str)

            df_order = df_order.fillna('')

            # 2. 获取当前系统中的数据
            df_current = self.data_manager.load_products()

            if df_current.empty:
                QMessageBox.warning(self, "提示", "系统中没有数据可排序")
                return

            # 3. 从排序文件中提取WB编号的顺序
            order_wb_col = None
            for col in df_order.columns:
                if 'WB编号' in col or 'WB货号' in col:
                    order_wb_col = col
                    break

            if order_wb_col is None:
                QMessageBox.warning(self, "错误", "排序文件中没有找到WB编号列")
                return

            ordered_wb_codes = []
            for idx, row in df_order.iterrows():
                wb = str(row[order_wb_col]).strip()
                if wb and wb != 'nan':
                    ordered_wb_codes.append(wb)

            if not ordered_wb_codes:
                QMessageBox.warning(self, "错误", "排序文件中没有有效的WB编号")
                return

            # 4. 按新顺序重新排列数据
            # 创建索引: WB编号 -> 行数据
            current_data_map = {}
            for idx, row in df_current.iterrows():
                wb = str(row.get('WB编号', '')).strip()
                if wb:
                    current_data_map[wb] = row

            # 按新顺序构建DataFrame
            new_rows = []
            matched_wb = []
            for wb in ordered_wb_codes:
                if wb in current_data_map:
                    new_rows.append(current_data_map[wb])
                    matched_wb.append(wb)
                else:
                    # 如果某个WB在系统中不存在，可以跳过或警告
                    pass

            # 如果有匹配不上的，也警告一下
            unmatched = set(ordered_wb_codes) - set(matched_wb)
            if unmatched:
                QMessageBox.warning(
                    self, "提示",
                    f"以下 {len(unmatched)} 个WB编号在当前系统中不存在，已跳过:\n{', '.join(list(unmatched)[:10])}"
                )

            if not new_rows:
                QMessageBox.warning(self, "错误", "排序文件中没有找到匹配的商品")
                return

            # 5. 保存新顺序
            df_sorted = pd.DataFrame(new_rows)
            self.data_manager.save_products(df_sorted)
            self.load_data()

            QMessageBox.information(
                self, "排序完成",
                f"✅ 已按新顺序排列 {len(new_rows)} 个商品\n"
                f"⚠️ 跳过系统中不存在的WB编号: {len(unmatched)} 个"
            )
        # 错误抛出
        except Exception as e:
            QMessageBox.critical(self, "错误", f"排序失败: {str(e)}")
            import traceback
            traceback.print_exc()

    # ===== 导出汇总表 =====
    def export_summary(self):
        """导出整个表格到桌面Excel（保留格式 + 条件格式）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.formatting.rule import Rule
            from openpyxl.styles.differential import DifferentialStyle

            df = self.table.df_full.copy()

            if df.empty:
                QMessageBox.information(self, "提示", "没有数据可导出")
                return

            store_name = self.store_name
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            filename = f"{store_name}汇总表.xlsx"
            filepath = os.path.join(desktop, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = "汇总表"

            headers = df.columns.tolist()

            # 识别数值列
            numeric_columns = ['售价', '库存'] + [col for col in headers if '促' in col]

            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )

            # 写入数据
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    col_name = headers[col_idx - 1]
                    cell = ws.cell(row=row_idx, column=col_idx)

                    if col_name in numeric_columns:
                        try:
                            if value and str(value).strip():
                                clean_value = str(value).replace(',', '').strip()
                                cell.value = float(clean_value)
                                if col_name == '售价':
                                    cell.number_format = '0.00'
                                else:
                                    cell.number_format = '0'
                            else:
                                cell.value = ''
                        except (ValueError, TypeError):
                            cell.value = value
                    else:
                        cell.value = value

                    cell.font = Font(name='微软雅黑', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin', color='CCCCCC'),
                        right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'),
                        bottom=Side(style='thin', color='CCCCCC')
                    )

            # ===== 条件格式 =====
            price_col_idx = None
            promo_cols = {}
            for idx, header in enumerate(headers, 1):
                if header == '售价':
                    price_col_idx = idx
                elif '促' in header and header not in ['促销']:
                    promo_cols[header] = idx

            if price_col_idx and promo_cols:
                price_col_letter = ws.cell(row=1, column=price_col_idx).column_letter
                last_row = len(df) + 1

                for promo_name, promo_col in promo_cols.items():
                    promo_col_letter = ws.cell(row=1, column=promo_col).column_letter
                    start_cell = f"{promo_col_letter}2"
                    end_cell = f"{promo_col_letter}{last_row}"
                    cells_range = f"{start_cell}:{end_cell}"

                    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    rule_green = Rule(
                        type='expression',
                        dxf=DifferentialStyle(fill=green_fill),
                        formula=[f'AND(${price_col_letter}2<={promo_col_letter}2, {promo_col_letter}2>0)'],
                        stopIfTrue=False
                    )
                    ws.conditional_formatting.add(cells_range, rule_green)

            # 自动调整列宽
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max(max_length + 2, 10), 40)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

            # 冻结首行
            ws.freeze_panes = 'A2'

            # 自动筛选
            last_col_letter = ws.cell(row=1, column=len(headers)).column_letter
            last_row = len(df) + 1
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

            wb.save(filepath)

            QMessageBox.information(
                self, "导出成功",
                f"已导出到桌面：\n{filename}\n\n"
            )
            self.status_label.setText(f"✅ 已导出: {filename}")

        except ImportError:
            QMessageBox.warning(self, "缺少依赖", "请安装 openpyxl：\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出失败：\n{str(e)}")
            self.status_label.setText("❌ 导出失败")

    def export_promo_adjustment(self):
        """导出促销调整数据（含变化数值和百分比）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.formatting.rule import Rule
            from openpyxl.styles.differential import DifferentialStyle

            df = self.table.df_full.copy()

            if df.empty:
                QMessageBox.information(self, "提示", "没有数据可导出")
                return

            promo_cols = [col for col in df.columns if '促' in col and col not in ['促销']]
            if not promo_cols:
                QMessageBox.information(self, "提示", "没有促销数据可导出")
                return

            if '售价' not in df.columns:
                QMessageBox.warning(self, "错误", "数据中没有售价列")
                return

            export_df = df.copy()
            export_df['售价'] = pd.to_numeric(export_df['售价'], errors='coerce')

            for promo in promo_cols:
                export_df[promo] = pd.to_numeric(export_df[promo], errors='coerce')
                diff_col = f"{promo}_变化数值"
                pct_col = f"{promo}_变化百分比"
                export_df[diff_col] = None
                export_df[pct_col] = None

                for idx, row in export_df.iterrows():
                    price = row.get('售价', 0)
                    promo_price = row.get(promo, None)
                    if price and price > 0 and promo_price and promo_price > 0:
                        diff = promo_price - price
                        pct = (diff / price) * 100
                        export_df.at[idx, diff_col] = round(diff, 2)
                        export_df.at[idx, pct_col] = round(pct, 1)

            base_cols = ['商品编号', 'WB编号', '类目', '库存', '仓库', '状态', '售价']
            export_cols = []
            for col in base_cols:
                if col in export_df.columns:
                    export_cols.append(col)
            for col in export_df.columns:
                if col not in base_cols and col not in promo_cols and not col.endswith('_变化数值') and not col.endswith('_变化百分比'):
                    export_cols.append(col)
            for promo in promo_cols:
                export_cols.append(promo)
                export_cols.append(f"{promo}_变化数值")
                export_cols.append(f"{promo}_变化百分比")

            final_df = export_df[export_cols]

            store_name = self.store_name
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            filename = f"{store_name}汇总表.xlsx"
            filepath = os.path.join(desktop, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = "汇总表"

            headers = final_df.columns.tolist()
            numeric_columns = ['售价', '库存'] + [col for col in headers if '促' in col and '变化' not in col]

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                if '变化' in header:
                    cell.fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )

            for row_idx, row in enumerate(final_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    col_name = headers[col_idx - 1]
                    cell = ws.cell(row=row_idx, column=col_idx)

                    if value is None:
                        cell.value = ''
                    elif col_name in numeric_columns:
                        try:
                            if value and str(value).strip() and str(value).lower() != 'nan':
                                clean_value = str(value).replace(',', '').strip()
                                cell.value = float(clean_value)
                                if '售价' in col_name:
                                    cell.number_format = '0.00'
                                else:
                                    cell.number_format = '0'
                            else:
                                cell.value = ''
                        except (ValueError, TypeError):
                            cell.value = value
                    else:
                        cell.value = value

                    cell.font = Font(name='微软雅黑', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin', color='CCCCCC'),
                        right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'),
                        bottom=Side(style='thin', color='CCCCCC')
                    )

            price_col_idx = None
            promo_col_idx_map = {}
            for idx, header in enumerate(headers, 1):
                if header == '售价':
                    price_col_idx = idx
                elif '促' in header and '变化' not in header:
                    promo_col_idx_map[header] = idx

            if price_col_idx and promo_col_idx_map:
                price_col_letter = ws.cell(row=1, column=price_col_idx).column_letter
                last_row = len(final_df) + 1

                for promo_name, promo_col in promo_col_idx_map.items():
                    promo_col_letter = ws.cell(row=1, column=promo_col).column_letter
                    start_cell = f"{promo_col_letter}2"
                    end_cell = f"{promo_col_letter}{last_row}"
                    cells_range = f"{start_cell}:{end_cell}"

                    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    rule_green = Rule(
                        type='expression',
                        dxf=DifferentialStyle(fill=green_fill),
                        formula=[f'AND(${price_col_letter}2<={promo_col_letter}2, {promo_col_letter}2>0)'],
                        stopIfTrue=False
                    )
                    ws.conditional_formatting.add(cells_range, rule_green)

            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max(max_length + 2, 12), 50)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

            ws.freeze_panes = 'A2'

            last_col_letter = ws.cell(row=1, column=len(headers)).column_letter
            last_row = len(final_df) + 1
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

            wb.save(filepath)

            QMessageBox.information(
                self, "导出成功",
                f"已导出到桌面：\n{filename}\n\n"
                f"✅ 每个促销列后增加了变化数值和变化百分比列\n"
                f"✅ 促销价绿色 = 售价 ≤ 促销价"
            )
            self.status_label.setText(f"✅ 已导出: {filename}")

        except ImportError:
            QMessageBox.warning(self, "缺少依赖", "请安装 openpyxl：\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出失败：\n{str(e)}")
            self.status_label.setText("❌ 导出失败")
            import traceback
            traceback.print_exc()

    def setup_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setSpacing(3)

        toolbar = QHBoxLayout()

        # ===== 统一按钮样式 =====
        btn_style_blue = """
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """
        btn_style_orange = """
            QPushButton {
                background-color: #ea580c;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #c2410c; }
        """
        btn_style_purple = """
            QPushButton {
                background-color: #7c3aed;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #6d28d9; }
        """
        btn_style_red = """
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #b91c1c; }
        """
        btn_style_green = """
            QPushButton {
                background-color: #0b8c5a;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #0a7048; }
        """
        btn_style_cyan = """
            QPushButton {
                background-color: #0891b2;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #0e7490; }
        """

        self.import_btn = QPushButton("📥 导入商品")
        self.import_btn.setStyleSheet(btn_style_blue)
        self.import_btn.clicked.connect(self.import_products)
        toolbar.addWidget(self.import_btn)

        self.add_product_btn = QPushButton("➕ 添加商品")
        self.add_product_btn.setStyleSheet(btn_style_blue)
        self.add_product_btn.clicked.connect(self.add_products_manually)
        toolbar.addWidget(self.add_product_btn)

        self.promo_btn = QPushButton("📊 上传促销")
        self.promo_btn.setStyleSheet(btn_style_orange)
        self.promo_btn.clicked.connect(self.upload_promotion)
        toolbar.addWidget(self.promo_btn)

        self.update_btn = QPushButton("🔄 更新库存/价格")
        self.update_btn.setStyleSheet(btn_style_blue)
        self.update_btn.clicked.connect(self.update_data)
        toolbar.addWidget(self.update_btn)

        self.import_competitor_btn = QPushButton("🔗 导入竞品链接")
        self.import_competitor_btn.setStyleSheet(btn_style_purple)
        self.import_competitor_btn.setToolTip("从Excel导入竞品链接\n第1列: WB编号\n第2~10列: 竞品链接(带批注价格)")
        self.import_competitor_btn.clicked.connect(self.import_competitors_from_template)
        toolbar.addWidget(self.import_competitor_btn)

        self.sync_all_img_btn = QPushButton("🖼️ 同步竞品图片")
        self.sync_all_img_btn.setStyleSheet(btn_style_purple)
        self.sync_all_img_btn.setToolTip("一键下载所有商品的竞品主图")
        self.sync_all_img_btn.clicked.connect(self.sync_all_competitor_images)
        toolbar.addWidget(self.sync_all_img_btn)

        toolbar.addStretch()

        self.refresh_btn = QPushButton("🔄 刷新数据")
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #2a7de1;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 20px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #1a5fb0; }
        """)
        self.refresh_btn.clicked.connect(self.refresh_data)
        toolbar.addWidget(self.refresh_btn)

        self.promo_combo = QComboBox()
        self.promo_combo.setPlaceholderText("选择要删除的促销")
        self.promo_combo.setMinimumWidth(180)
        self.promo_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #d0d9e8;
                border-radius: 6px;
                padding: 6px 12px;
                background-color: white;
                font-size: 13px;
            }
            QComboBox:hover { border-color: #2a7de1; }
            QComboBox::drop-down { border: none; }
        """)
        toolbar.addWidget(self.promo_combo)

        self.delete_promo_btn = QPushButton("🗑️ 删除促销")
        self.delete_promo_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #b91c1c; }
            QPushButton:disabled { background-color: #9ca3af; }
        """)
        self.delete_promo_btn.clicked.connect(self.delete_promo_from_toolbar)
        self.delete_promo_btn.setEnabled(False)
        toolbar.addWidget(self.delete_promo_btn)

        self.refresh_promo_btn = QPushButton("🔄 刷新促销")
        self.refresh_promo_btn.setStyleSheet("""
            QPushButton {
                background-color: #0b8c5a;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #0a7048; }
            QPushButton:disabled { background-color: #9ca3af; }
        """)
        self.refresh_promo_btn.setToolTip("清理过期促销并重新排序")
        self.refresh_promo_btn.clicked.connect(self.refresh_promotions)
        self.refresh_promo_btn.setEnabled(False)
        toolbar.addWidget(self.refresh_promo_btn)

        self.toggle_adj_btn = QPushButton("📊 商品导出")
        self.toggle_adj_btn.setStyleSheet("""
            QPushButton {
                background-color: #6b21a5;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #581c87; }
        """)
        self.toggle_adj_btn.setToolTip("导出当前表格数据到桌面Excel")
        self.toggle_adj_btn.clicked.connect(self.export_summary)
        toolbar.addWidget(self.toggle_adj_btn)

        self.promo_adj_btn = QPushButton("📊 促销调整")
        self.promo_adj_btn.setStyleSheet("""
            QPushButton {
                background-color: #d97706;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover { background-color: #b45309; }
        """)
        self.promo_adj_btn.setToolTip("导出促销调整数据（含变化数值和百分比）")
        self.promo_adj_btn.clicked.connect(self.export_promo_adjustment)
        toolbar.addWidget(self.promo_adj_btn)

        self.clear_btn = QPushButton("🗑️ 清空汇总表")
        self.clear_btn.setStyleSheet(btn_style_red)
        self.clear_btn.clicked.connect(self.clear_data)
        toolbar.addWidget(self.clear_btn)

        main_layout.addLayout(toolbar)

        self.filter_container = QWidget()
        self.filter_container.setStyleSheet("""
            QWidget {
                background-color: #f0f2f5;
                border: 1px solid #dce4ec;
                border-bottom: none;
            }
        """)
        self.filter_container.setFixedHeight(32)
        self.filter_layout = QHBoxLayout(self.filter_container)
        self.filter_layout.setContentsMargins(2, 2, 2, 2)
        self.filter_layout.setSpacing(0)
        main_layout.addWidget(self.filter_container)

        self.table = PromoTableWidget()
        self.table.delete_requested.connect(self.delete_product)
        main_layout.addWidget(self.table)

        status_layout = QHBoxLayout()
        self.status_label = QLabel("就绪")
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()
        main_layout.addLayout(status_layout)

        self.setLayout(main_layout)

    # ===== 刷新促销 =====
    def refresh_promotions(self):
        """刷新促销：清理过期 + 重新排序"""
        try:
            self.status_label.setText("🔄 正在刷新促销...")
            QApplication.processEvents()
            self.promo_manager.clean_expired()
            self.promo_manager.reorder_promotions()
            self.load_data()
            promo_count = len(self.promo_manager.get_promo_columns())
            QMessageBox.information(self, "完成", f"✅ 促销已刷新\n当前共 {promo_count} 个促销")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"刷新促销失败:\n{str(e)}")
            self.status_label.setText("❌ 刷新促销失败")

    # ===== 同步竞品图片 =====
    def sync_all_competitor_images(self):
        """一键同步所有商品的竞品图片"""
        import urllib.request

        df = self.data_manager.load_products()
        print(f"📊 加载到 {len(df)} 个商品")

        if df.empty:
            QMessageBox.information(self, "提示", "暂无商品数据")
            return

        all_tasks = []
        for _, row in df.iterrows():
            product_wb = str(row.get('WB编号', '')).strip()
            print(f"🔍 处理商品: {product_wb}")
            if not product_wb or product_wb == 'nan':
                continue
            manager = CompetitorManager(self.store_name, product_wb)
            data = manager.get_all_competitors()
            print(f"   └─ 竞品数量: {len(data)}")
            for wb_code in data.keys():
                img_path = manager.get_image_path(wb_code)
                if img_path and os.path.exists(img_path):
                    print(f"      ⏭️ 跳过已有图片: {wb_code}")
                    continue
                else:
                    print(f"      📥 需要下载: {wb_code}")
                all_tasks.append((product_wb, wb_code))

        print(f"📋 最终需要下载的任务数: {len(all_tasks)}")

        if not all_tasks:
            QMessageBox.information(self, "提示", "所有竞品已有图片，无需同步")
            return

        reply = QMessageBox.question(
            self, "同步竞品图片",
            f"共 {len(all_tasks)} 个竞品需要下载主图，是否继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            print("❌ 用户取消同步")
            return

        self._sync_progress_dialog = QDialog(self)
        self._sync_progress_dialog.setWindowTitle("同步竞品图片")
        self._sync_progress_dialog.setFixedSize(420, 120)
        layout = QVBoxLayout(self._sync_progress_dialog)
        self._sync_label = QLabel(f"正在下载 0/{len(all_tasks)}...")
        layout.addWidget(self._sync_label)
        self._sync_bar = QProgressBar()
        self._sync_bar.setMaximum(len(all_tasks))
        layout.addWidget(self._sync_bar)
        self._sync_detail = QLabel("")
        self._sync_detail.setStyleSheet("font-size: 11px; color: #888;")
        layout.addWidget(self._sync_detail)
        self._sync_progress_dialog.show()

        class ImgThread(QThread):
            progress = pyqtSignal(int, str, str)
            finished = pyqtSignal(int, int)

            def __init__(self, tasks, store_name):
                super().__init__()
                self.tasks = tasks
                self.store_name = store_name

            def run(self):
                import urllib.request
                success = 0
                fail = 0
                for i, (product_wb, wb_code) in enumerate(self.tasks):
                    downloaded = False
                    vol = wb_code[:4]
                    part = wb_code[:6]

                    # ===== 修改：basket 从 01 到 60，只保留两种域名，注释掉 .cn =====
                    url_configs = []
                    for b in range(1, 61):
                        url_configs.append(
                            f"https://basket-{b:02d}.wildberries.ru/vol{vol}/part{part}/{wb_code}/images/big/1.webp")
                    for b in range(1, 61):
                        url_configs.append(
                            f"https://basket-{b:02d}.wbbasket.ru/vol{vol}/part{part}/{wb_code}/images/big/1.webp")
                    # ===== 注释掉 .cn 域名 =====
                    # for b in range(50, 9, -1):
                    #     url_configs.append(
                    #         f"https://basket-{b:02d}.wildberries.cn/vol{vol}/part{part}/{wb_code}/images/big/1.webp")

                    for url in url_configs:
                        try:
                            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                            with urllib.request.urlopen(req, timeout=5) as resp:
                                img_data = resp.read()
                            if img_data and len(img_data) > 1000:
                                manager = CompetitorManager(self.store_name, product_wb)
                                manager.update_image(wb_code, img_data)
                                success += 1
                                downloaded = True
                                self.progress.emit(i + 1, f"✅ {wb_code}", "")
                                break
                        except:
                            continue

                    if not downloaded:
                        fail += 1
                        self.progress.emit(i + 1, f"❌ {wb_code}", "")
                self.finished.emit(success, fail)

        self._sync_thread = ImgThread(all_tasks, self.store_name)
        self._sync_thread.progress.connect(self._on_sync_progress)
        self._sync_thread.finished.connect(self._on_sync_all_done)
        self._sync_thread.start()

    def _on_sync_progress(self, cur, status, detail):
        self._sync_label.setText(f"{status}  ({cur}/{self._sync_bar.maximum()})")
        self._sync_bar.setValue(cur)

    def _on_sync_all_done(self, success, fail):
        self._sync_progress_dialog.close()
        self.status_label.setText(f"✅ 同步完成: {success}成功/{fail}失败")

    # ===== 导入竞品链接 =====
    def import_competitors_from_template(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择包含竞品链接的Excel文件", self.default_dir,
            "Excel文件 (*.xlsx *.xls)"
        )
        if not file_path:
            return

        excel = None
        try:
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(os.path.abspath(file_path))
            ws = wb.ActiveSheet

            total_imported = 0
            total_prices = 0
            skipped = 0
            row = 2

            while True:
                product_wb_val = ws.Cells(row, 1).Value
                if product_wb_val is None:
                    break

                if isinstance(product_wb_val, float) and product_wb_val == int(product_wb_val):
                    product_wb = str(int(product_wb_val))
                else:
                    product_wb = str(product_wb_val).strip()

                if product_wb:
                    manager = CompetitorManager(self.store_name, product_wb)

                    for col in range(2, 11):
                        cell = ws.Cells(row, col)
                        cell_value = str(cell.Value).strip() if cell.Value else ""
                        if not cell_value:
                            continue

                        wb_code = self._extract_wb_from_url(cell_value)
                        if not wb_code:
                            skipped += 1
                            continue

                        if manager.add_competitor(wb_code):
                            total_imported += 1

                        try:
                            comment = cell.Comment
                            if comment:
                                comment_text = comment.Text()
                                if comment_text:
                                    price = self._extract_price_from_comment(comment_text)
                                    if price is not None:
                                        manager.update_price(wb_code, price)
                                        total_prices += 1
                        except:
                            pass

                row += 1
                if row > 10000:
                    break

            wb.Close(False)
            excel.Quit()
            excel = None

            msg = f"✅ 导入竞品: {total_imported} 个\n💰 设置价格: {total_prices} 个\n⏭️ 跳过无效链接: {skipped} 个"
            QMessageBox.information(self, "导入完成", msg)
            self.status_label.setText(f"✅ 竞品导入完成: {total_imported} 个")

        except ImportError:
            QMessageBox.critical(self, "缺少依赖", "请先安装: pip install pywin32")
        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入竞品链接时出错:\n{str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            if excel:
                try:
                    excel.Quit()
                except:
                    pass

    def _extract_wb_from_url(self, url: str) -> str:
        pattern = r'/catalog/(\d+)/detail\.aspx'
        match = re.search(pattern, url)
        return match.group(1) if match else ""

    def _extract_price_from_comment(self, comment_text: str):
        if not comment_text:
            return None
        lines = comment_text.strip().split('\n')
        for line in reversed(lines):
            numbers = re.findall(r'\d+\.?\d*', line.strip())
            if numbers:
                try:
                    return float(numbers[-1])
                except:
                    continue
        return None

    # ===== 添加商品 =====
    def add_products_manually(self):
        from PyQt6.QtWidgets import QTextEdit, QDialogButtonBox

        dialog = QDialog(self)
        dialog.setWindowTitle("添加商品")
        dialog.setMinimumSize(400, 300)

        layout = QVBoxLayout(dialog)

        label = QLabel("请输入商品编号，每行一个：")
        layout.addWidget(label)

        text_edit = QTextEdit()
        text_edit.setPlaceholderText("G07412429274721\nG07402072401676\n...")
        layout.addWidget(text_edit)

        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        raw_text = text_edit.toPlainText().strip()
        if not raw_text:
            QMessageBox.warning(self, "提示", "未输入任何商品编号")
            return

        lines = [line.strip() for line in raw_text.split('\n') if line.strip()]

        df = self.data_manager.load_products()
        existing = set(df['商品编号'].astype(str).str.strip().tolist())

        added = 0
        skipped = 0
        for code in lines:
            if code in existing:
                skipped += 1
                continue
            new_row = pd.DataFrame([{
                '商品编号': code,
                'WB编号': '',
                '类目': '',
                '库存': 0,
                '仓库': 'FBW',
                '状态': '正常',
                '售价': 0
            }])
            df = pd.concat([df, new_row], ignore_index=True)
            existing.add(code)
            added += 1

        self.data_manager.save_products(df)
        self.load_data()

        msg = f"✅ 添加: {added} 个"
        if skipped:
            msg += f"\n⏭️ 跳过重复: {skipped} 个"
        QMessageBox.information(self, "完成", msg)

    # ===== 搜索 =====
    def on_search_text_changed(self, text):
        self._pending_search_text = text.strip()
        self._search_timer.stop()
        self._search_timer.start(300)

    def _do_search(self):
        text = self._pending_search_text
        try:
            self.clear_search_highlight()
            self.search_results = []
            self.search_current_index = -1
            if not text:
                return
            search_text = text.lower()
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    try:
                        item = self.table.item(row, col)
                        item_text = ""
                        if item is not None:
                            item_text = item.text()
                        if not item_text:
                            widget = self.table.cellWidget(row, col)
                            if widget:
                                label = widget.findChild(QLabel)
                                if label:
                                    item_text = label.text()
                        if item_text and search_text in item_text.lower():
                            self.search_results.append((row, col))
                    except:
                        continue
            if self.search_results:
                self.search_current_index = 0
                QTimer.singleShot(10, lambda: self.highlight_search_result(self.search_results[0]))
        except:
            pass

    def highlight_search_result(self, position):
        try:
            row, col = position
            self.clear_search_highlight()
            self.search_highlighted_row = row
            for c in range(self.table.columnCount()):
                item = self.table.item(row, c)
                if item and item.text():
                    item.setBackground(QColor(173, 216, 230))
            first_item = self.table.item(row, 0)
            if first_item:
                self.table.scrollToItem(first_item, QAbstractItemView.ScrollHint.PositionAtCenter)
            self.table.selectRow(row)
            self.table.setFocus()
        except:
            pass

    def clear_search_highlight(self):
        try:
            if self.search_highlighted_row >= 0:
                row = self.search_highlighted_row
                for c in range(self.table.columnCount()):
                    item = self.table.item(row, c)
                    if item and item.text():
                        key = (row, c)
                        if key in self.table.row_bg_cache:
                            item.setBackground(self.table.row_bg_cache[key])
                        else:
                            item.setBackground(QColor(255, 255, 255))
                self.search_highlighted_row = -1
        except:
            self.search_highlighted_row = -1

    def find_next_search(self):
        if not self.search_results or self.search_current_index < 0:
            return
        self.search_current_index = (self.search_current_index + 1) % len(self.search_results)
        self.highlight_search_result(self.search_results[self.search_current_index])

    def find_prev_search(self):
        if not self.search_results or self.search_current_index < 0:
            return
        self.search_current_index = (self.search_current_index - 1) % len(self.search_results)
        self.highlight_search_result(self.search_results[self.search_current_index])

    def copy_selected_cells(self):
        try:
            selected_ranges = self.table.selectedRanges()
            if not selected_ranges:
                if self.copy_status:
                    self.copy_status.setText("⚠️ 未选中")
                    QTimer.singleShot(2000, lambda: self.copy_status.setText(""))
                return
            all_rows = []
            for range_obj in selected_ranges:
                for row in range(range_obj.topRow(), range_obj.bottomRow() + 1):
                    row_data = []
                    for col in range(range_obj.leftColumn(), range_obj.rightColumn() + 1):
                        item = self.table.item(row, col)
                        if item is not None and item.text():
                            row_data.append(item.text())
                        else:
                            widget = self.table.cellWidget(row, col)
                            if widget:
                                label = widget.findChild(QLabel)
                                if label and label.text():
                                    row_data.append(label.text())
                                else:
                                    row_data.append('')
                            else:
                                row_data.append('')
                    if row_data:
                        all_rows.append('\t'.join(row_data))
            if all_rows:
                QApplication.clipboard().setText('\n'.join(all_rows))
                if self.copy_status:
                    self.copy_status.setText(f"✅ {len(all_rows)}行")
                    QTimer.singleShot(2000, lambda: self.copy_status.setText(""))
        except:
            pass

    def clear_filters(self):
        for col, widget in self.filter_widgets.items():
            if isinstance(widget, MultiSelectComboBox):
                widget.reset()
            elif isinstance(widget, QComboBox):
                widget.setCurrentIndex(0)
        if self.color_filter_widget:
            self.color_filter_widget.setCurrentIndex(0)
        if self.search_input:
            self.search_input.clear()
        if self.category_input:
            self.category_input.clear()
        self.clear_search_highlight()
        self.table.df_filtered = self.table.df_full.copy()
        self.table.load_filtered_data(self.table.df_full)
        self.status_label.setText(f"✅ 已清空筛选，共 {len(self.table.df_full)} 个商品")

    def setup_filter_row(self, df):
        while self.filter_layout.count():
            item = self.filter_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        self.filter_widgets = {}
        self.color_filter_widget = None
        self.category_input = None
        self.copy_btn = None
        self.copy_status = None
        self.search_input = None

        all_columns = self.table.get_all_columns()
        normal_filter_columns = ['仓库', '库存', '状态']

        for col_index, col in enumerate(all_columns):
            width = self.table.columnWidth(col_index)

            if col == '序号':
                label = QLabel("序号")
                label.setStyleSheet("font-weight: bold; color: #888; font-size: 10px;")
                label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                label.setFixedWidth(width)
                label.setFixedHeight(24)
                self.filter_layout.addWidget(label)
                continue

            if col == '类目':
                self.category_input = QLineEdit()
                self.category_input.setPlaceholderText("输入类目筛选...")
                self.category_input.setStyleSheet("""
                    QLineEdit {
                        border: 1px solid #ccc;
                        border-radius: 3px;
                        padding: 1px 4px;
                        font-size: 10px;
                        background: white;
                    }
                    QLineEdit:focus { border: 1px solid #0078d7; }
                """)
                self.category_input.setFixedWidth(width)
                self.category_input.setFixedHeight(24)
                self.category_input.textChanged.connect(self.on_filter_changed)
                self.filter_layout.addWidget(self.category_input)
                continue

            if col in normal_filter_columns:
                if col in ['仓库', '状态']:
                    # 多选筛选
                    if col in df.columns:
                        unique_values = df[col].dropna().unique().tolist()
                        unique_values = [str(v) for v in unique_values if str(v).strip()]
                        unique_values.sort()
                        multi = MultiSelectComboBox(unique_values)
                    else:
                        multi = MultiSelectComboBox([])
                    multi.setFixedWidth(width)
                    multi.setFixedHeight(24)
                    multi.selection_changed.connect(lambda: self.on_filter_changed())
                    multi.setProperty("column", col)
                    multi.setProperty("filter_type", "normal")
                    self.filter_widgets[col] = multi
                    self.filter_layout.addWidget(multi)
                else:
                    # 库存等普通下拉
                    combo = QComboBox()
                    combo.addItem("(全部)")
                    combo.setStyleSheet("""
                        QComboBox {
                            border: 1px solid #ccc;
                            border-radius: 3px;
                            padding: 1px 2px;
                            font-size: 10px;
                            background: white;
                        }
                        QComboBox:focus { border: 1px solid #0078d7; }
                    """)
                    combo.currentTextChanged.connect(lambda: self.on_filter_changed())
                    combo.setProperty("column", col)
                    combo.setProperty("filter_type", "normal")
                    combo.setFixedWidth(width)
                    combo.setFixedHeight(24)
                    if col in df.columns:
                        unique_values = df[col].dropna().unique().tolist()
                        unique_values = [str(v) for v in unique_values if str(v).strip()]
                        try:
                            unique_values.sort(key=lambda x: int(float(x)))
                        except:
                            unique_values.sort()
                        for val in unique_values:
                            combo.addItem(val)
                    self.filter_widgets[col] = combo
                    self.filter_layout.addWidget(combo)

            elif col in self.table.promo_columns:
                # 促销列 - 显示促销表头名称作为默认文本
                combo = QComboBox()
                combo.setStyleSheet("""
                    QComboBox {
                        border: 1px solid #ccc;
                        border-radius: 3px;
                        padding: 1px 2px;
                        font-size: 10px;
                        background: white;
                    }
                    QComboBox:focus { border: 1px solid #0078d7; }
                """)

                # 添加选项：第一个选项显示促销表头名称
                combo.addItem(col)  # 直接显示列名，如"促1"
                combo.addItem("✅ 有促销")
                combo.addItem("❌ 无促销")

                combo.currentTextChanged.connect(lambda: self.on_filter_changed())
                combo.setProperty("column", col)
                combo.setProperty("filter_type", "promo")
                combo.setFixedWidth(width)
                combo.setFixedHeight(24)

                # 设置默认选中第一个选项（促销表头名称）
                combo.setCurrentIndex(0)

                self.filter_widgets[col] = combo
                self.filter_layout.addWidget(combo)

            elif col == '商品编号':
                # 搜索框 + 复制按钮 + 状态标签，总宽度 = width
                search_w = width - 74  # 复制按钮28 + 状态标签46

                self.search_input = QLineEdit()
                self.search_input.setPlaceholderText("🔍 搜索...")
                self.search_input.setStyleSheet("""
                    QLineEdit {
                        border: 1px solid #ccc;
                        border-radius: 3px;
                        padding: 1px 4px;
                        font-size: 10px;
                        background: white;
                    }
                    QLineEdit:focus { border: 1px solid #0078d7; }
                """)
                self.search_input.setFixedWidth(search_w)
                self.search_input.setFixedHeight(24)
                self.search_input.textChanged.connect(self.on_search_text_changed)
                self.search_input.returnPressed.connect(self.find_next_search)
                self.filter_layout.addWidget(self.search_input)

                self.copy_btn = QPushButton("📋")
                self.copy_btn.setToolTip("选中单元格后点击复制")
                self.copy_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                self.copy_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e8f0fe;
                        color: #1a73e8;
                        border: 1px solid #1a73e8;
                        border-radius: 3px;
                        padding: 1px 4px;
                        font-size: 10px;
                        font-weight: 500;
                    }
                    QPushButton:hover { background-color: #d2e3fc; }
                """)
                self.copy_btn.clicked.connect(self.copy_selected_cells)
                self.copy_btn.setFixedWidth(28)
                self.copy_btn.setFixedHeight(24)
                self.filter_layout.addWidget(self.copy_btn)

                self.copy_status = QLabel("")
                self.copy_status.setStyleSheet(
                    "color: #5a7a9a; font-size: 10px; background: transparent; border: none;")
                self.copy_status.setFixedWidth(46)
                self.filter_layout.addWidget(self.copy_status)

            else:
                placeholder = QLabel("")
                placeholder.setFixedWidth(width)
                placeholder.setFixedHeight(24)
                self.filter_layout.addWidget(placeholder)

        color_label = QLabel("🎨")
        color_label.setStyleSheet("font-size: 12px;")
        color_label.setFixedWidth(20)
        color_label.setFixedHeight(24)
        self.filter_layout.addWidget(color_label)

        self.color_filter_widget = QComboBox()
        self.color_filter_widget.addItems(["(全部)", "🟢 绿色高亮", "⬜ 正常颜色"])
        self.color_filter_widget.setStyleSheet("""
            QComboBox {
                border: 1px solid #ccc;
                border-radius: 3px;
                padding: 1px 2px;
                font-size: 10px;
                background: white;
            }
            QComboBox:focus { border: 1px solid #0078d7; }
        """)
        self.color_filter_widget.setFixedWidth(100)
        self.color_filter_widget.setFixedHeight(24)
        self.color_filter_widget.currentTextChanged.connect(lambda: self.on_filter_changed())
        self.filter_layout.addWidget(self.color_filter_widget)
        self.filter_layout.addStretch()

    def on_filter_changed(self):
        filters = {}
        for col, widget in self.filter_widgets.items():
            if isinstance(widget, MultiSelectComboBox):
                selected = widget.get_selected()
                filter_type = widget.property("filter_type")
                if filter_type == "normal" and selected:
                    filters[col] = selected
            elif isinstance(widget, QComboBox):
                text = widget.currentText()
                filter_type = widget.property("filter_type")
                if filter_type == "normal" and text and text != "(全部)":
                    filters[col] = text
                elif filter_type == "promo":
                    if text == "✅ 有促销":
                        filters[col] = "__HAS_VALUE__"
                    elif text == "❌ 无促销":
                        filters[col] = "__EMPTY__"

        color_filter = None
        if self.color_filter_widget:
            ct = self.color_filter_widget.currentText()
            if ct == "🟢 绿色高亮":
                color_filter = "green"
            elif ct == "⬜ 正常颜色":
                color_filter = "white"

        category_text = self.category_input.text().strip() if self.category_input else ""

        self.table.apply_filter_with_category(filters, color_filter, category_text)

        row_count = self.table.rowCount()
        total_count = len(self.table.df_full)
        desc_parts = []
        for k, v in filters.items():
            if v == "__HAS_VALUE__":
                desc_parts.append(f"{k}=有促销")
            elif v == "__EMPTY__":
                desc_parts.append(f"{k}=无促销")
            elif isinstance(v, list):
                desc_parts.append(f"{k}=[{','.join(v)}]")
            else:
                desc_parts.append(f"{k}={v}")
        if color_filter:
            desc_parts.append(f"颜色={color_filter}")
        if category_text:
            desc_parts.append(f"类目≈{category_text}")
        if desc_parts:
            self.status_label.setText(f"🔍 {', '.join(desc_parts)} | {row_count}/{total_count}")
        else:
            self.status_label.setText(f"✅ 共 {total_count} 个商品")

    def load_data(self):
        try:
            df = self.data_manager.load_products()
            promo_columns = self.promo_manager.get_promo_columns()
            self.table.set_store_name(self.store_name)

            df = self.promo_manager.apply_promo_to_df(df)

            self.promo_combo.clear()
            if promo_columns:
                self.promo_combo.addItem("-- 全部促销 --")
                self.promo_combo.addItems(promo_columns)
                self.delete_promo_btn.setEnabled(True)
                self.refresh_promo_btn.setEnabled(True)
            else:
                self.promo_combo.addItem("-- 全部促销 --")
                self.delete_promo_btn.setEnabled(False)
                self.refresh_promo_btn.setEnabled(False)
            meta = {col: self.promo_manager.get_promo_meta(col) for col in promo_columns if
                    self.promo_manager.get_promo_meta(col)}
            self.table.set_promo_meta(meta)
            self.table.load_data(df, promo_columns)
            self.setup_filter_row(df)
            self.status_label.setText(f"✅ 共 {len(df)} 个商品，促销列: {len(promo_columns)}")
        except Exception as e:
            self.status_label.setText(f"❌ 加载失败: {str(e)}")

    def refresh_data(self):
        try:
            self.status_label.setText("🔄 正在刷新...")
            QApplication.processEvents()
            self.data_manager = DataManager(self.store_name)
            self.promo_manager = PromotionManager(self.data_manager)
            self.load_data()
            self.status_label.setText("✅ 刷新完成")
        except Exception as e:
            self.status_label.setText("❌ 刷新失败")

    def delete_promo_from_toolbar(self):
        promo_name = self.promo_combo.currentText()
        if not promo_name:
            return
        if promo_name == "-- 全部促销 --":
            if QMessageBox.question(self, "确认", "删除所有促销？",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
                self.promo_manager.delete_all_promotions()
                self.load_data()
            return
        if QMessageBox.question(self, "确认", f"删除「{promo_name}」？",
                                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            self.promo_manager.delete_promotion(promo_name)
            self.load_data()

    def delete_product(self, wb_code: str, product_code: str):
        df = self.data_manager.load_products()
        df['WB编号'] = df['WB编号'].astype(str).str.strip()
        mask = df['WB编号'] == wb_code
        if mask.any():
            df = df[~mask]
            self.data_manager.save_products(df)
            self.load_data()
            self.status_label.setText(f"🗑️ 已删除商品: {product_code} (WB: {wb_code})")
        else:
            QMessageBox.warning(self, "错误", f"未找到商品: {product_code}")

    def import_products(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择商品数据文件", "", "Excel/CSV文件 (*.xlsx *.xls *.csv)")
        if not file_path:
            return
        # ✅ 增加 "排序数据" 选项
        mode, ok = QInputDialog.getItem(
            self, "导入模式",
            "请选择导入方式：",
            ["追加数据", "覆盖数据", "排序数据"],
            0, False
        )
        if not ok or mode == "取消":
            return

        if mode == "排序数据":
            self.import_and_sort(file_path)
        else:
            mode = "append" if mode == "追加数据" else "overwrite"
            success, skipped, errors = self.data_manager.import_products(file_path, mode)
            QMessageBox.information(self, "导入结果", f"成功: {success}, 跳过: {skipped}")
            self.load_data()

    def upload_promotion(self):
        dialog = UploadDialog("上传促销", show_date=True, default_dir=self.default_dir, parent=self)
        if dialog.exec():
            if dialog.file_path:
                start_date, end_date = dialog.get_dates()
                self.promo_manager.import_promotion(dialog.file_path, start_date, end_date)
                self.load_data()

    def update_data(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择更新模板文件", self.default_dir,
            "Excel/CSV文件 (*.xlsx *.xls *.csv)"
        )
        if file_path:
            self.data_manager.update_from_template(file_path)
            self.load_data()

    def clear_data(self):
        if QMessageBox.question(self, "确认", "清空所有数据？",
                                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            self.data_manager.save_products(self.data_manager._create_empty_df())
            self.promo_manager.promotions = []
            self.promo_manager._save()
            self.load_data()