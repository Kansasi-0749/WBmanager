"""Excel 导出：汇总表与促销调整表共用一个写入器。"""
from pathlib import Path

import pandas as pd

from . import config


def export_summary(df: pd.DataFrame, filepath) -> str:
    """导出汇总表（保留原条件格式与样式）。"""
    return _write_excel(df, filepath, adjustment_mode=False)


def export_promo_adjustment(df: pd.DataFrame, filepath) -> str:
    """导出促销调整表（每个促销列后追加变化数值/百分比列）。"""
    promo_cols = [col for col in df.columns if "促" in col and col not in ["促销"]]
    export_df = df.copy()
    export_df["售价"] = pd.to_numeric(export_df["售价"], errors="coerce")

    for promo in promo_cols:
        export_df[promo] = pd.to_numeric(export_df[promo], errors="coerce")
        diff_col = f"{promo}_变化数值"
        pct_col = f"{promo}_变化百分比"
        export_df[diff_col] = None
        export_df[pct_col] = None

        price = export_df["售价"]
        promo_price = export_df[promo]
        valid = price.notna() & promo_price.notna() & (price > 0) & (promo_price > 0)
        diff = promo_price - price
        pct = (diff / price) * 100
        export_df.loc[valid, diff_col] = diff[valid].round(2)
        export_df.loc[valid, pct_col] = pct[valid].round(1)

    base_cols = config.BASE_COLUMNS
    export_cols = [c for c in base_cols if c in export_df.columns]
    export_cols += [
        c
        for c in export_df.columns
        if c not in base_cols
        and c not in promo_cols
        and not c.endswith("_变化数值")
        and not c.endswith("_变化百分比")
    ]
    for promo in promo_cols:
        export_cols.append(promo)
        export_cols.append(f"{promo}_变化数值")
        export_cols.append(f"{promo}_变化百分比")

    return _write_excel(export_df[export_cols], filepath, adjustment_mode=True)


def _write_excel(df: pd.DataFrame, filepath, *, adjustment_mode: bool = False) -> str:
    """构建带样式、条件格式、冻结首行与自动筛选的 Excel 工作簿。"""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.styles.differential import DifferentialStyle

    headers = list(df.columns)

    if adjustment_mode:
        numeric_columns = ["售价", "库存"] + [
            c for c in headers if "促" in c and "变化" not in c
        ]
        min_width, max_width = 12, 50
    else:
        numeric_columns = ["售价", "库存"] + [c for c in headers if "促" in c]
        min_width, max_width = 10, 40

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总表"

    header_fill_normal = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill_change = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    thin_black = Border(*[Side(style="thin", color="000000")] * 4)
    thin_gray = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        if adjustment_mode and "变化" in header:
            cell.fill = header_fill_change
        else:
            cell.fill = header_fill_normal
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_black

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            col_name = headers[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx)

            if value is None:
                cell.value = ""
            elif col_name in numeric_columns:
                try:
                    if value and str(value).strip() and str(value).lower() != "nan":
                        clean_value = str(value).replace(",", "").strip()
                        cell.value = float(clean_value)
                        cell.number_format = "0.00" if "售价" in col_name else "0"
                    else:
                        cell.value = ""
                except (ValueError, TypeError):
                    cell.value = value
            else:
                cell.value = value

            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_gray

    # 条件格式：售价 ≤ 促销价 且促销价 > 0 时绿色高亮
    price_col_idx = None
    promo_cols = {}
    for idx, header in enumerate(headers, 1):
        if header == "售价":
            price_col_idx = idx
        elif "促" in header and header not in ["促销"] and "变化" not in header:
            promo_cols[header] = idx

    if price_col_idx and promo_cols:
        price_col_letter = ws.cell(row=1, column=price_col_idx).column_letter
        last_row = len(df) + 1
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        for promo_name, promo_col in promo_cols.items():
            promo_col_letter = ws.cell(row=1, column=promo_col).column_letter
            cells_range = f"{promo_col_letter}2:{promo_col_letter}{last_row}"
            rule_green = Rule(
                type="expression",
                dxf=DifferentialStyle(fill=green_fill),
                formula=[f"AND(${price_col_letter}2<={promo_col_letter}2, {promo_col_letter}2>0)"],
                stopIfTrue=False,
            )
            ws.conditional_formatting.add(cells_range, rule_green)

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

    ws.freeze_panes = "A2"
    last_col_letter = ws.cell(row=1, column=len(headers)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(df) + 1}"

    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    return str(filepath)
